from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter

folder = Path(__file__).parent     # Folder containing all your Excel files
output = "ICICI_NAV_Master.xlsx"

dfs = []

for file in folder.glob("*.xlsx"):
    df = pd.read_excel(file)
    dfs.append(df)

master = pd.concat(dfs, ignore_index=True)

# Optional: Sort by fund name and date
master.sort_values(["fund_name", "date"], inplace=True)

master.to_excel(output, index=False)

# ---------- Formatting ----------
wb = load_workbook(output)
ws = wb.active

# Freeze header
ws.freeze_panes = "A2"

# Add filters
ws.auto_filter.ref = ws.dimensions

# Thick border after each fund
thick = Side(style="thick")

for r in range(2, ws.max_row):
    if r == ws.max_row or ws[f"B{r}"].value != ws[f"B{r+1}"].value:
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).border = Border(bottom=thick)

# Auto-fit columns
for col in ws.columns:
    max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
    ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 40)

wb.save(output)

print("Master Excel created successfully!")